import pandas as pd
import numpy as np
from llama_index.core import Document, VectorStoreIndex
# from llama_index.embeddings.openai import OpenAIEmbedding # Assuming OpenAI embeddings
# from llama_index.core import Settings # For setting global embedding models
import openpyxl # For reading Excel files
import io # For handling file streams if necessary

# --- Enhanced Styling Instructions for Plotly ---
enhanced_styling_instructions = """
# Comprehensive Plotly Styling Guide for Auto-Analyst Agents

## 1. General Principles:
   - **Clarity and Readability:** Prioritize clear communication of insights. Ensure legible fonts, sufficient contrast, and uncluttered layouts.
   - **Consistency:** Maintain a consistent visual style across multiple charts in a report or dashboard.
   - **Purpose-Driven Design:** Choose chart types and styling elements that best serve the analytical goal.
   - **Interactivity:** Leverage Plotly's interactivity (hover, zoom, pan) to allow users to explore data. Use `st.plotly_chart(fig, use_container_width=True)` in Streamlit.
   - **Accessibility:** Consider color-blind friendly palettes and clear text alternatives where possible.

## 2. Color Palettes:
   - **Sequential Data:** Use for ordered data that progresses from low to high.
     - Examples: `px.colors.sequential.Viridis`, `px.colors.sequential.Blues`, `px.colors.sequential.Greens`.
   - **Diverging Data:** Use when the data has a meaningful midpoint (e.g., zero) and deviates in two directions.
     - Examples: `px.colors.diverging.RdBu`, `px.colors.diverging.Picnic`, `px.colors.diverging.Spectral`.
   - **Qualitative/Categorical Data:** Use for distinct categories without inherent order. Ensure colors are distinguishable.
     - Examples: `px.colors.qualitative.Plotly`, `px.colors.qualitative.Safe`, `px.colors.qualitative.D3`.
     - For many categories, consider grouping or using a base color with varying shades/tints.
   - **Custom Palettes:** Define as `color_discrete_map={'Category1': 'blue', 'Category2': 'red'}` in `px` functions.
   - **Color Blindness:** Test palettes using simulators. `px.colors.qualitative.Plotly` is generally good.

## 3. Chart-Specific Styling Templates:

   ### a. Bar Charts (`px.bar` or `go.Bar`):
      - **Layout:** `fig.update_layout(title_text='Your Title', xaxis_title='X-axis Label', yaxis_title='Y-axis Label', barmode='group'/'stack')`
      - **Colors:** `color='category_column'` for grouped bars, `color_discrete_sequence=['#1f77b4', '#ff7f0e', ...]` for custom sequence.
      - **Text on Bars:** `text_auto=True` or `fig.update_traces(texttemplate='%{text:.2s}', textposition='outside')`
      - **Hover:** `hover_data={'column_name': True}` or customize `hovertemplate`.
      - **Example:**
        ```python
        # fig = px.bar(df, x='x_col', y='y_col', color='group_col', title='Bar Chart Title',
        #              labels={'x_col':'X Axis', 'y_col':'Y Axis'},
        #              text_auto=True, barmode='group')
        # fig.update_layout(legend_title_text='Group Legend')
        ```

   ### b. Line Charts (`px.line` or `go.Scatter` with `mode='lines'`):
      - **Layout:** `fig.update_layout(title_text='Trend Over Time', xaxis_title='Date/Time', yaxis_title='Value')`
      - **Markers:** `fig.update_traces(mode='lines+markers')` or `px.line(..., markers=True)`.
      - **Multiple Lines:** `color='category_column'` or add multiple `go.Scatter` traces.
      - **Line Style:** `fig.update_traces(line=dict(dash='dash'/'dot', width=2, color='red'))` for specific traces.
      - **Example:**
        ```python
        # fig = px.line(df, x='date_col', y='value_col', color='series_col', title='Time Series Trend',
        #               markers=True, labels={'value_col': 'Metric Value'})
        # fig.update_xaxes(rangeslider_visible=True) # Optional: add a range slider for time series
        ```

   ### c. Scatter Plots (`px.scatter` or `go.Scatter` with `mode='markers'`):
      - **Layout:** `fig.update_layout(title_text='Relationship Between X and Y', xaxis_title='X Variable', yaxis_title='Y Variable')`
      - **Size & Color Encoding:** `size='size_column'`, `color='color_column'`.
      - **Trendlines:** `px.scatter(..., trendline='ols'/'lowess')` for Ordinary Least Squares or Locally Weighted Scatterplot Smoothing.
      - **Hover:** Customize `hovertemplate` to show relevant info.
      - **Example:**
        ```python
        # fig = px.scatter(df, x='x_var', y='y_var', color='category', size='magnitude',
        #                  title='Scatter Plot with Size and Color Encoding',
        #                  trendline='ols', hover_name='identifier_col')
        ```

   ### d. Histograms (`px.histogram` or `go.Histogram`):
      - **Layout:** `fig.update_layout(title_text='Distribution of Feature', xaxis_title='Feature Value', yaxis_title='Frequency/Density')`
      - **Bins:** `nbins=30` or `xbins=dict(start=0, end=100, size=5)`.
      - **Normalization:** `histnorm='probability density'` or `histnorm='percent'`.
      - **Overlaid Histograms:** `barmode='overlay'`, then `fig.update_traces(opacity=0.75)`.
      - **Example:**
        ```python
        # fig = px.histogram(df, x='numeric_feature', color='group_col', marginal='box', # or 'rug', 'violin'
        #                    title='Histogram with Marginal Box Plot', nbins=50, opacity=0.7)
        ```

   ### e. Pie Charts (`px.pie` or `go.Pie`):
      - **Layout:** `fig.update_layout(title_text='Composition of Categories')`
      - **Values & Names:** `values='numeric_column'`, `names='category_column'`.
      - **Hole (Donut Chart):** `hole=0.4`.
      - **Text Info:** `fig.update_traces(textinfo='percent+label', pull=[0.1, 0, 0])` to explode a slice.
      - **Caution:** Avoid pie charts for >5-7 categories or when comparing similar-sized slices. Bar charts are often better.
      - **Example:**
        ```python
        # fig = px.pie(df, values='amount_col', names='category_col', title='Category Proportions', hole=0.3)
        # fig.update_traces(textposition='inside', textinfo='percent+label')
        ```

   ### f. Box Plots (`px.box` or `go.Box`):
      - **Layout:** `fig.update_layout(title_text='Distribution Comparison by Group', yaxis_title='Value')`
      - **Orientation:** `orientation='h'` for horizontal.
      - **Points:** `points='all'` (shows all data points), `'outliers'`, or `False`.
      - **Notched Box:** `notched=True` for confidence interval around median.
      - **Example:**
        ```python
        # fig = px.box(df, x='category_col', y='value_col', color='group_col',
        #                title='Box Plot Comparison', points='outliers', notched=True)
        ```

   ### g. Heatmaps (`px.imshow` or `go.Heatmap`):
      - **Data:** Typically a 2D matrix (e.g., correlation matrix).
      - **Color Scale:** `color_continuous_scale='RdBu_r'` (for correlations), `'Viridis'`.
      - **Annotations:** `text_auto=True` to show values on cells. `fig.update_traces(texttemplate="%{z:.2f}")`.
      - **Example (Correlation Matrix):**
        ```python
        # corr_matrix = df_numeric.corr()
        # fig = px.imshow(corr_matrix, text_auto=True, aspect="auto",
        #                 color_continuous_scale='RdBu_r', zmin=-1, zmax=1,
        #                 title='Correlation Heatmap')
        ```

   ### h. 3D Scatter/Surface Plots (`px.scatter_3d`, `go.Scatter3d`, `go.Surface`):
      - Use sparingly, can be hard to interpret.
      - Ensure clear labeling and good default camera angle.
      - `fig.update_layout(scene=dict(xaxis_title='X', yaxis_title='Y', zaxis_title='Z'))`
      - **Example:**
        ```python
        # fig = px.scatter_3d(df, x='x_col', y='y_col', z='z_col', color='color_col', size='size_col',
        #                     title='3D Scatter Plot')
        # fig.update_layout(margin=dict(l=0, r=0, b=0, t=40))
        ```

## 4. Layout and Annotations:
   - **Titles & Labels:** Always include descriptive titles and axis labels. `fig.update_layout(title_font_size=20, xaxis_title_font_size=16, ...)`
   - **Legends:** `fig.update_layout(legend_title_text='Legend', legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1))`
   - **Margins:** `fig.update_layout(margin=dict(l=50, r=50, t=80, b=50))` (left, right, top, bottom).
   - **Annotations:** `fig.add_annotation(x=..., y=..., text="Important Point", showarrow=True, arrowhead=1)`
   - **Shapes (lines, rectangles):** `fig.add_shape(type="line", x0=..., y0=..., x1=..., y1=..., line=dict(color="Red", dash="dash"))`

## 5. Handling Large Datasets:
   - **Sampling:** If `len(df) > 50000`, consider `df_sample = df.sample(n=50000, random_state=42)` for scatter plots.
   - **Aggregation:** For bar/line charts, ensure data is aggregated appropriately before plotting (e.g., `df.groupby('date').mean()`).
   - **Datashader (Advanced):** For very large scatter plots, Plotly can integrate with Datashader to render aggregated representations. This is more complex.
     ```python
     # import datashader as ds
     # import datashader.transfer_functions as tf
     # from colorcet import fire
     # cvs = ds.Canvas(plot_width=400, plot_height=400)
     # agg = cvs.points(df, 'x_col', 'y_col')
     # img = tf.shade(agg, cmap=fire)[::-1].to_pil()
     # fig.add_layout_image(dict(source=img, xref="x", yref="y", x=df['x_col'].min(), y=df['y_col'].max(), sizex=..., sizey=...))
     ```
     (Agent should simplify this or state it's using an aggregated view if directly implementing).

## 6. Streamlit Specifics:
   - Always use `st.plotly_chart(fig, use_container_width=True)` to display the chart.
   - Do not use `fig.show()` as it's for standalone environments.
   - If generating multiple related charts, consider using `st.tabs` or `st.columns`.

## 7. Code Structure for Agents:
   - Import necessary libraries: `import plotly.express as px`, `import plotly.graph_objects as go`.
   - Prepare data (select columns, aggregate if needed).
   - Create the figure: `fig = px.chart_type(...)` or `fig = go.Figure(...)`.
   - Update layout and traces: `fig.update_layout(...)`, `fig.update_traces(...)`.
   - The final line of code for visualization should be `st.plotly_chart(fig, use_container_width=True)`.

By following these guidelines, agents can produce informative, aesthetically pleasing, and effective visualizations.
"""

# --- Helper Functions ---
def correct_num(num_str):
    """Cleans a string to be a valid number (float) by removing commas and percent signs."""
    if isinstance(num_str, (int, float)):
        return num_str
    if isinstance(num_str, str):
        num_str = num_str.replace(',', '').replace('%', '').strip()
        try:
            return float(num_str)
        except ValueError:
            return np.nan # Return NaN if conversion fails
    return np.nan

def return_vals(df, col):
    """
    Provides a summary of a column: data type, number of unique values,
    and top N most frequent values with their counts, missing values, and basic stats.
    """
    col_summary = {}
    col_summary['dtype'] = str(df[col].dtype)
    col_summary['unique_values_count'] = df[col].nunique()

    # Handle potential errors with value_counts, especially for unhashable types
    try:
        top_n = df[col].value_counts(dropna=False).nlargest(5) # Include NaNs in counts
        col_summary['top_n_values'] = {str(k): v for k, v in top_n.items()}
    except TypeError:
        # If direct value_counts fails (e.g., list in cells), try converting to string first
        try:
            top_n = df[col].astype(str).value_counts(dropna=False).nlargest(5)
            col_summary['top_n_values'] = {str(k): v for k, v in top_n.items()}
            col_summary['note'] = "Converted to string for value counts due to unhashable type."
        except Exception:
            col_summary['top_n_values'] = "Could not compute top N values (possibly unhashable types)."

    col_summary['missing_values_count'] = df[col].isnull().sum()
    col_summary['missing_values_percent'] = (df[col].isnull().sum() / len(df) * 100) if len(df) > 0 else 0

    if pd.api.types.is_numeric_dtype(df[col]):
        # Ensure there are non-NaN values before calculating stats
        if df[col].notna().any():
            col_summary['mean'] = df[col].mean()
            col_summary['median'] = df[col].median()
            col_summary['std_dev'] = df[col].std()
            col_summary['min'] = df[col].min()
            col_summary['max'] = df[col].max()
            col_summary['q1'] = df[col].quantile(0.25)
            col_summary['q3'] = df[col].quantile(0.75)
        else:
            col_summary['mean'] = col_summary['median'] = col_summary['std_dev'] = \
            col_summary['min'] = col_summary['max'] = col_summary['q1'] = col_summary['q3'] = np.nan

    elif pd.api.types.is_datetime64_any_dtype(df[col]):
        if df[col].notna().any():
            col_summary['min_date'] = df[col].min()
            col_summary['max_date'] = df[col].max()
            col_summary['time_span_days'] = (df[col].max() - df[col].min()).days if pd.notnull(df[col].min()) and pd.notnull(df[col].max()) else None
        else:
            col_summary['min_date'] = col_summary['max_date'] = col_summary['time_span_days'] = None
    return col_summary

def generate_excel_overview(excel_file_source):
    """
    Generates an overview of an Excel file, including sheet names,
    number of rows/columns per sheet, and common columns between sheets.
    Accepts file path or file-like object.
    """
    xls = None
    try:
        xls = pd.ExcelFile(excel_file_source)
    except Exception as e:
        return f"Error opening Excel file: {str(e)}\n"

    sheet_names = xls.sheet_names
    if not sheet_names:
        return "Excel File Overview:\n  No sheets found in the Excel file.\n"

    overview = f"Excel File Overview:\n  File contains {len(sheet_names)} sheet(s): {', '.join(sheet_names)}\n\n"
    
    sheet_details = []
    sheet_columns_map = {} # To store columns for relationship inference

    # Determine if excel_file_source is a path or stream for openpyxl
    is_path = isinstance(excel_file_source, str)

    for sheet_name in sheet_names:
        try:
            # Read only a few rows for initial schema (faster)
            df_sheet_sample = xls.parse(sheet_name, nrows=5) 
            
            actual_rows = "N/A (full parse needed for exact count)"
            actual_cols = len(df_sheet_sample.columns) # From sample

            # Try to get actual dimensions using openpyxl if it's a file path
            if is_path:
                try:
                    workbook = openpyxl.load_workbook(excel_file_source, read_only=True, data_only=True)
                    if sheet_name in workbook.sheetnames:
                        sheet_obj = workbook[sheet_name]
                        actual_rows = sheet_obj.max_row
                        actual_cols = sheet_obj.max_column # openpyxl might give more accurate column count
                    workbook.close()
                except Exception: # Fallback if openpyxl fails
                    pass # Keep default values

            sheet_info = f"  Sheet '{sheet_name}':\n"
            sheet_info += f"    - Approx. Dimensions (from sample): {len(df_sheet_sample)} rows (sampled) x {len(df_sheet_sample.columns)} columns\n"
            if is_path: # Only show actual dimensions if we could read them
                 sheet_info += f"    - Actual Dimensions (from metadata): {actual_rows} rows x {actual_cols} columns\n"
            sheet_info += f"    - Columns: {', '.join(df_sheet_sample.columns.tolist())}\n"
            sheet_details.append(sheet_info)
            sheet_columns_map[sheet_name] = set(df_sheet_sample.columns.tolist())
        except Exception as e:
            sheet_details.append(f"  Sheet '{sheet_name}': Error reading sheet sample - {str(e)}\n")
            sheet_columns_map[sheet_name] = set()
            
    overview += "\n".join(sheet_details)
    
    # Infer relationships (common columns)
    if len(sheet_names) > 1:
        overview += "\nPotential Relationships (Common Columns):\n"
        relationships_found = False
        for i in range(len(sheet_names)):
            for j in range(i + 1, len(sheet_names)):
                s1_name = sheet_names[i]
                s2_name = sheet_names[j]
                if s1_name in sheet_columns_map and s2_name in sheet_columns_map:
                    common = list(sheet_columns_map[s1_name].intersection(sheet_columns_map[s2_name]))
                    if common:
                        overview += f"  - Sheets '{s1_name}' and '{s2_name}' share columns: {', '.join(common)}\n"
                        relationships_found = True
        if not relationships_found:
            overview += "  - No obvious common columns found between sheets based on headers of sampled data.\n"
            
    return overview

def make_data(data_source, dataset_name="Dataset", target_sheet_name=None):
    """
    Enhanced function to generate a detailed schema description for a dataset.
    Supports pandas DataFrame, CSV file path/stream, or Excel file path/stream.
    For Excel, if target_sheet_name is None, provides an overview of all sheets
    and schema for the first sheet. If target_sheet_name is specified, provides
    schema for that sheet.
    """
    schema_description = f"Schema for {dataset_name}:\n"
    df_to_describe = None
    excel_source_name = ""

    if hasattr(data_source, 'name') and isinstance(data_source.name, str): # Check if it's a file stream with a name
        excel_source_name = data_source.name
    elif isinstance(data_source, str): # Check if it's a file path
        excel_source_name = data_source


    if isinstance(data_source, pd.DataFrame):
        df_to_describe = data_source
        schema_description += f"  Source: DataFrame in memory.\n"
    elif (isinstance(data_source, str) and (data_source.lower().endswith('.csv'))) or \
         (hasattr(data_source, 'name') and isinstance(data_source.name, str) and data_source.name.lower().endswith('.csv')):
        try:
            if hasattr(data_source, 'seek') and callable(data_source.seek): data_source.seek(0)
            df_to_describe = pd.read_csv(data_source)
            schema_description += f"  Source: CSV file ('{excel_source_name or 'uploaded_file.csv'}').\n"
        except Exception as e:
            return schema_description + f"  Error reading CSV: {str(e)}"

    elif (isinstance(data_source, str) and (data_source.lower().endswith(('.xlsx', '.xls', '.xlsm')))) or \
         (hasattr(data_source, 'name') and isinstance(data_source.name, str) and data_source.name.lower().endswith(('.xlsx', '.xls', '.xlsm'))):
        
        try:
            if hasattr(data_source, 'seek') and callable(data_source.seek): data_source.seek(0)
            # Pass the original source (path or stream) to generate_excel_overview
            excel_overview = generate_excel_overview(data_source) 
            schema_description += excel_overview
            
            if hasattr(data_source, 'seek') and callable(data_source.seek): data_source.seek(0) # Reset stream again for parsing
            xls = pd.ExcelFile(data_source)
            if not xls.sheet_names:
                return schema_description + "  Error: No sheets found in the Excel file."

            sheet_to_load = None
            if target_sheet_name and target_sheet_name in xls.sheet_names:
                sheet_to_load = target_sheet_name
                schema_description += f"\nDetailed Schema for Sheet '{target_sheet_name}':\n"
            else:
                sheet_to_load = xls.sheet_names[0]
                schema_description += f"\nDetailed Schema for First Sheet ('{sheet_to_load}'):\n"
                if target_sheet_name: # If a sheet was specified but not found
                     schema_description += f"  (Note: Requested sheet '{target_sheet_name}' not found, showing first sheet instead.)\n"
            
            df_to_describe = xls.parse(sheet_to_load)

        except Exception as e:
            return schema_description + f"  Error processing Excel file: {str(e)}"
            
    elif hasattr(data_source, 'read'): # Generic file stream, try CSV then Excel
        if hasattr(data_source, 'seek') and callable(data_source.seek): data_source.seek(0)
        try:
            df_to_describe = pd.read_csv(data_source)
            schema_description += f"  Source: Uploaded file stream (parsed as CSV).\n"
        except Exception:
            try: # Try Excel if CSV fails
                if hasattr(data_source, 'seek') and callable(data_source.seek): data_source.seek(0)
                excel_overview = generate_excel_overview(data_source)
                schema_description += excel_overview
                
                if hasattr(data_source, 'seek') and callable(data_source.seek): data_source.seek(0)
                xls = pd.ExcelFile(data_source) # This might fail if it's not excel
                if not xls.sheet_names: return schema_description + "  No sheets found."

                sheet_to_load = xls.sheet_names[0] # Default to first sheet
                schema_description += f"\nDetailed Schema for First Sheet ('{sheet_to_load}'):\n"
                df_to_describe = xls.parse(sheet_to_load)
                schema_description = schema_description.replace("(parsed as CSV)", "(parsed as Excel)")
            except Exception as e_excel:
                 return schema_description + f"  Error: Could not parse uploaded file stream as CSV or Excel. Details: {str(e_excel)}"
    else:
        return "Error: Unsupported data source type. Please provide a pandas DataFrame, CSV/Excel file path, or file stream."

    if df_to_describe is None:
        # If we only generated an Excel overview and didn't load a sheet, that's fine.
        # Otherwise, it's an error.
        if "Detailed Schema for Sheet" not in schema_description and "Excel File Overview" in schema_description:
             return schema_description # Return just the overview
        return schema_description + "  Error: Could not load data for schema generation."


    schema_description += f"  - Number of rows: {len(df_to_describe)}\n"
    schema_description += f"  - Number of columns: {len(df_to_describe.columns)}\n"
    if len(df_to_describe) == 0:
        schema_description += "  - Note: The selected dataset/sheet is empty.\n"
        return schema_description

    schema_description += "Column Details:\n"

    for col in df_to_describe.columns:
        summary = return_vals(df_to_describe, col)
        schema_description += f"  - Column '{col}':\n"
        schema_description += f"    - Data Type (Pandas): {summary['dtype']}\n"
        
        inferred_type = "Unknown"
        col_data = df_to_describe[col] # Use the actual column data for inference
        if pd.api.types.is_numeric_dtype(col_data): inferred_type = "Numeric"
        elif pd.api.types.is_datetime64_any_dtype(col_data): inferred_type = "Datetime"
        elif pd.api.types.is_timedelta64_dtype(col_data): inferred_type = "Timedelta"
        elif pd.api.types.is_bool_dtype(col_data): inferred_type = "Boolean"
        elif pd.api.types.is_categorical_dtype(col_data): inferred_type = "Categorical"
        elif pd.api.types.is_string_dtype(col_data) or col_data.dtype == 'object':
            # Further check if object column is mostly strings or mixed
            non_null_sample = col_data.dropna().unique()[:5]
            if all(isinstance(x, str) for x in non_null_sample) and non_null_sample.size > 0:
                inferred_type = "Text/String"
            elif non_null_sample.size == 0 and col_data.isnull().all(): # All nulls
                 inferred_type = "Object (All Nulls)"
            else:
                inferred_type = "Object/Mixed"
        
        schema_description += f"    - Inferred General Type: {inferred_type}\n"
        schema_description += f"    - Unique Values: {summary['unique_values_count']}\n"
        schema_description += f"    - Missing Values: {summary['missing_values_count']} ({summary['missing_values_percent']:.2f}%)\n"
        
        if 'note' in summary: schema_description += f"    - Note: {summary['note']}\n"

        if summary['unique_values_count'] <= 10 and summary['unique_values_count'] > 0 :
            schema_description += f"    - Top Values (Counts): {summary['top_n_values']}\n"
        elif summary['unique_values_count'] > 0:
            sample_unique_values = [str(v) for v in col_data.dropna().unique()[:3]]
            schema_description += f"    - Sample Unique Values: [{', '.join(sample_unique_values)} ...]\n"

        if pd.api.types.is_numeric_dtype(col_data) and col_data.notna().any():
            schema_description += f"    - Statistics: Mean={summary.get('mean', np.nan):.2f}, Median={summary.get('median', np.nan):.2f}, StdDev={summary.get('std_dev', np.nan):.2f}\n"
            schema_description += f"                  Min={summary.get('min', np.nan):.2f}, Max={summary.get('max', np.nan):.2f}, Q1={summary.get('q1', np.nan):.2f}, Q3={summary.get('q3', np.nan):.2f}\n"
        elif pd.api.types.is_datetime64_any_dtype(col_data) and col_data.notna().any():
            schema_description += f"    - Date Range: From {summary.get('min_date', 'N/A')} to {summary.get('max_date', 'N/A')}\n"
            if summary.get('time_span_days') is not None:
                 schema_description += f"    - Time Span: {summary['time_span_days']} days\n"
    
    schema_description += "\nGeneral Dataset Summary:\n"
    duplicate_rows = df_to_describe.duplicated().sum()
    schema_description += f"  - Contains {duplicate_rows} duplicate rows ({ (duplicate_rows/len(df_to_describe)*100) if len(df_to_describe)>0 else 0 :.2f}%).\n"
    
    potential_id_cols = [
        col for col in df_to_describe.columns 
        if df_to_describe[col].nunique() == len(df_to_describe) and df_to_describe[col].isnull().sum() == 0
    ]
    if potential_id_cols:
        schema_description += f"  - Potential ID/Key Column(s) (all unique, no missing): {', '.join(potential_id_cols)}\n"

    return schema_description


def initiatlize_retrievers(data_description_str, styling_instructions_str, embedding_model=None):
    """
    Initializes LlamaIndex retrievers for data description and styling instructions.
    Relies on global LlamaIndex Settings for embedding model if `embedding_model` is None.
    """
    retrievers = {}
    data_doc = Document(text=data_description_str)
    style_doc = Document(text=styling_instructions_str)

    # If embedding_model is passed, it's used directly. Otherwise, LlamaIndex defaults or global Settings apply.
    # Ensure LlamaIndex's Settings.embed_model is configured if not passing embedding_model.
    # e.g., from llama_index.core import Settings; from llama_index.embeddings.openai import OpenAIEmbedding
    # Settings.embed_model = OpenAIEmbedding() # Requires OPENAI_API_KEY in env

    if embedding_model:
        data_index = VectorStoreIndex.from_documents([data_doc], embed_model=embedding_model)
        style_index = VectorStoreIndex.from_documents([style_doc], embed_model=embedding_model)
    else:
        data_index = VectorStoreIndex.from_documents([data_doc]) # Uses global/default embed_model
        style_index = VectorStoreIndex.from_documents([style_doc]) # Uses global/default embed_model
        
    retrievers['dataframe_index_retriever'] = data_index.as_retriever(similarity_top_k=3) # Get more schema context
    retrievers['style_index_retriever'] = style_index.as_retriever(similarity_top_k=2) # Get more styling context

    return retrievers

styling_instructions = enhanced_styling_instructions

if __name__ == '__main__':
    print("--- Testing CSV File ---")
    csv_data_content = """ID,Name,Age,Salary,Department,JoinDate,Rating,Comments
1,Alice,25,"50,000",HR,2020-01-15,4.5,Good worker
2,Bob,30,"60,000.50",Engineering,2019-07-22,4.0,Excellent coder
3,Charlie,35,75000,HR,2021-03-10,4.8,
4,David,28,"52,000",Marketing,2020-11-05,3.9,Creative
5,Eve,22,48000.00,Engineering,2022-05-30,4.2,Promising
1,Alice,25,"50,000",HR,2020-01-15,4.5,Good worker
"""
    csv_stream = io.StringIO(csv_data_content)
    csv_schema = make_data(csv_stream, dataset_name="Dummy CSV Data from Stream")
    print(csv_schema)
    print("\n" + "="*50 + "\n")

    print("--- Testing Excel File ---")
    excel_io = io.BytesIO()
    with pd.ExcelWriter(excel_io, engine='openpyxl') as writer:
        pd.DataFrame({
            'EmployeeID': [1, 2, 3, 1], 'Name': ['Alice', 'Bob', 'Charlie', 'Alice'], 'Salary': [50000, 60000, 75000, 50000]
        }).to_excel(writer, sheet_name='Employees', index=False)
        pd.DataFrame({
            'ProjectID': ['P101', 'P102', 'P103'], 'ProjectName': ['Alpha', 'Beta', 'Gamma'], 'LeadEmployeeID': [1, 2, 1]
        }).to_excel(writer, sheet_name='Projects', index=False)
        pd.DataFrame().to_excel(writer, sheet_name='EmptySheet', index=False)
    excel_io.seek(0) # Reset stream position

    excel_schema_overview = make_data(excel_io, dataset_name="Dummy Excel Workbook from Stream")
    print(excel_schema_overview)
    print("\n" + "="*50 + "\n")
    
    excel_io.seek(0)
    excel_schema_sheet1 = make_data(excel_io, dataset_name="Dummy Excel Workbook from Stream", target_sheet_name='Employees')
    print(excel_schema_sheet1)
    print("\n" + "="*50 + "\n")

    print("--- Testing DataFrame Input ---")
    df_in_memory = pd.DataFrame({
        'A': [1,2,None,4,5,6,7,8,9,10,11,12], # More data for stats
        'B': ['x','y','z','x','y','z','x','y','z','x','y','z'],
        'C': pd.to_datetime(['2023-01-01', '2023-01-02', '2023-01-03'] * 4),
        'D': [True, False, True] * 4,
        'E': [1.1, 2.2, 3.3, 4.4, 5.5, 6.6, 7.7, 8.8, 9.9, 10.1, 11.1, 12.1]
    })
    df_schema = make_data(df_in_memory, dataset_name="In-Memory DataFrame")
    print(df_schema)

    # To run retriever initialization test:
    # 1. Ensure OPENAI_API_KEY is in your environment.
    # 2. Install llama-index-embeddings-openai: pip install llama-index-embeddings-openai
    # 3. Uncomment the following:
    # print("\n" + "="*50 + "\n")
    # print("--- Testing Retriever Initialization (requires OpenAI API Key & LlamaIndex setup) ---")
    # try:
    #     from llama_index.core import Settings
    #     from llama_index.embeddings.openai import OpenAIEmbedding
    #     Settings.embed_model = OpenAIEmbedding() # This will use OPENAI_API_KEY from env
    #
    #     retrievers_instance = initiatlize_retrievers(df_schema, enhanced_styling_instructions)
    #     print(f"Retriever for DataFrame Index created: {retrievers_instance['dataframe_index_retriever'] is not None}")
    #     print(f"Retriever for Style Index created: {retrievers_instance['style_index_retriever'] is not None}")
    #
    #     # Example query
    #     if retrievers_instance['dataframe_index_retriever']:
    #         relevant_schema_parts = retrievers_instance['dataframe_index_retriever'].retrieve("details of column A")
    #         if relevant_schema_parts:
    #             print(f"\nRetrieved schema for 'column A':\n{relevant_schema_parts[0].get_content()[:300]}...")
    #         else:
    #             print("No relevant schema parts found for 'column A'.")
    # except Exception as e:
    #     print(f"Could not initialize retrievers: {e}")
    #     print("This is expected if OpenAI API key is not set or llama_index libraries are not fully configured.")
